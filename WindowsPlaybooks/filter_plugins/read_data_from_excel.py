from ansible.errors import AnsibleError
import json

from collections import OrderedDict
import openpyxl
from itertools import islice
import os
import re


class FilterModule(object):

    def filter_records_by_switch_names(self,column_value,excel_file_path):
        ''''
         Read the excel file using pandas package and filter the records based on the column_value
        '''

        if not os.path.exists(excel_file_path):
            raise AnsibleError("Excel file at path %s does not exist"%excel_file_path)

        try:
            wb_obj = openpyxl.load_workbook(filename=excel_file_path)
            #Take the default sheetshee
            sheet = wb_obj.worksheets[0]
            rows = []
            # Iterate through each row in worksheet and fetch values into dict
            for row in islice(sheet.values, 1, sheet.max_row):
                entry = OrderedDict()
                if column_value in row[0]:
                    entry['switch_names'] = row[0]
                    entry['acl_num'] = row[1]
                    entry['acl_type'] = row[2]
                    entry['source'] = row[3]
                    entry['src_mask_bits'] = row[4]
                    entry['destination'] = row[5]
                    entry['dest_mask_bits'] = row[6]
                    entry['ports'] = row[7]
                    entry['operation'] = row[8]
                    rows.append(entry)

        except Exception as e:
            raise AnsibleError(message=e.message)

        return json.dumps(rows)

    def get_all_user_records(self,excel_file_path):
        ''''
         Read the excel file using pandas package and filter the records based on the column_value
        '''

        if not os.path.exists(excel_file_path):
            raise AnsibleError("Excel file at path %s does not exist"%excel_file_path)

        try:
            wb_obj = openpyxl.load_workbook(filename=excel_file_path)
            #Take the default sheet
            sheet = wb_obj.worksheets[0]
            rows = []
            # Iterate through each row in worksheet and fetch values into dict
            for row in islice(sheet.values, 1, sheet.max_row):
                entry = OrderedDict()
                entry['user_name'] = row[0]
                entry['full_name'] = row[1]
                entry['description'] = row[2]
                entry['password'] = row[3]
                entry['can_user_change_password'] = row[4]
                entry['operation'] = row[5]
                entry['servers']=row[6]
                rows.append(entry)

        except Exception as e:
            raise AnsibleError(message=e)

        return json.dumps(rows)

    def generate_dest_ports(self,ports_information):
        port_list=[]
        #splits the ports_information by newline
        individual_ports_infos=ports_information.split('\n')
        for ports in individual_ports_infos:

            # ignore the empty lines
            if not ports.strip() == '':
            # now splits the ports by forward slash
                try:
                    port_type=ports.strip().split('/')[0]
                    port_number=int(ports.strip().split('/')[1])
                except Exception as e:
                    raise AnsibleError(message="Unable to parse the port details %s. "
                                           "There must be single port entry per line.Example: tcp/60, then in new line it must another "
                                           "entry. "%ports)
                port_info={}
                port_info['port_type'] = port_type
                port_info['port_num'] = port_number
                #now append to the port list
                port_list.append(port_info)

        #Verify the port list should not empty
        if len(port_list)==0:
            raise AnsibleError(message="The port list is empty.It is mandatory for extended configuration")

        return port_list

    def validate_ports(self,ports_information):
        port_list = []
        #splits the ports_information by newline
        individual_ports_infos=ports_information.split('\n')
        for ports in individual_ports_infos:
            #ignore the empty lines
            if not ports.strip()=='':
                # now splits the ports by forward slash
                try:
                    port_type=ports.strip().split('/')[0]
                    port_number=int(ports.strip().split('/')[1])
                except Exception as e:
                    raise AnsibleError(message="Unable to parse the port details %s. "
                                           "There must be single port entry per line.Example: tcp/60, then in new line it must another "
                                           "entry. "%ports)
                port_list.append(ports)

        # Verify the port list should not empty
        if len(port_list) == 0:
            raise AnsibleError(message="The port list is empty.It is mandatory for extended configuration")

        return port_list

    def validate_ip(self,ip_address):
        try:
            result = [0 <= int(x) < 256 for x in
                      re.split('\.', re.match(r'^\d+\.\d+\.\d+\.\d+$', ip_address.strip()).group(0))].count(True) == 4
        except Exception as e:
            raise AnsibleError("Invalid IP address %s provided."%ip_address)

        if not result:
            raise AnsibleError("Invalid IP address %s provided."%ip_address)

        return result


    def filters(self):
        return {
            'filter_records_by_switch_names': self.filter_records_by_switch_names,
            'generate_dest_ports':self.generate_dest_ports,
            'get_all_user_records':self.get_all_user_records,
            'validate_ports': self.validate_ports,
            'validate_ip': self.validate_ip
        }




